from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, abort
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, date
from functools import wraps
from models import db, User, Solicitud, Comentario, Bitacora, Documento, TemasSolicitud, RutaTransporte, SolicitudIngeniero, cdmx_now
import os, uuid, csv, io, zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Response

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'soluciones-logisticas-secret-2026')
# Render usa 'postgres://' — psycopg3 requiere 'postgresql+psycopg://'
_db_url = os.environ.get('DATABASE_URL', 'sqlite:///database.db')
if _db_url.startswith('postgres://'):
    _db_url = _db_url.replace('postgres://', 'postgresql+psycopg://', 1)
elif _db_url.startswith('postgresql://'):
    _db_url = _db_url.replace('postgresql://', 'postgresql+psycopg://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 280,
    'pool_timeout': 20,
    'pool_size': 5,
    'max_overflow': 2,
}
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
UPLOAD_BASE = os.path.join(os.path.dirname(__file__), 'uploads')
ALLOWED_EXTENSIONS = {'pdf','xlsx','xls','pptx','docx','png','jpg','jpeg','zip'}

db.init_app(app)



login_manager = LoginManager(app)
login_manager.login_view = 'login'

# ── Inicialización diferida de BD (lazy) ─────────────────────────────────
# Evita que db.create_all() bloquee el arranque si PostgreSQL está dormido
_db_initialized = False

@app.before_request
def _lazy_init_db():
    global _db_initialized
    if _db_initialized:
        return
    try:
        with db.engine.connect() as conn:
            from sqlalchemy import text
            migraciones = [
                "ALTER TABLE solicitudes ADD COLUMN IF NOT EXISTS subtipo VARCHAR(10)",
                "ALTER TABLE solicitudes ADD COLUMN IF NOT EXISTS prioridad_sugerida INTEGER",
                "ALTER TABLE solicitudes ADD COLUMN IF NOT EXISTS prioridad_comercial INTEGER",
                "ALTER TABLE solicitudes ADD COLUMN IF NOT EXISTS prioridad_estado VARCHAR(20) DEFAULT 'pendiente'",
                "ALTER TABLE solicitudes ADD COLUMN IF NOT EXISTS solicitud_origen_id INTEGER",
                "ALTER TABLE solicitudes ADD COLUMN IF NOT EXISTS fecha_info_completa TIMESTAMP",
                "ALTER TABLE solicitudes ADD COLUMN IF NOT EXISTS dias_analisis INTEGER",
                """CREATE TABLE IF NOT EXISTS solicitud_ingenieros (
                    id SERIAL PRIMARY KEY,
                    solicitud_id INTEGER REFERENCES solicitudes(id),
                    ingeniero_id INTEGER REFERENCES users(id),
                    es_principal BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT NOW()
                )""",
                """CREATE TABLE IF NOT EXISTS rutas_transporte (
                    id SERIAL PRIMARY KEY,
                    solicitud_id INTEGER REFERENCES solicitudes(id),
                    orden INTEGER DEFAULT 1,
                    origen_estado VARCHAR(100) NOT NULL DEFAULT '',
                    origen_ciudad VARCHAR(100) NOT NULL DEFAULT '',
                    destino_estado VARCHAR(100) NOT NULL DEFAULT '',
                    destino_ciudad VARCHAR(100) NOT NULL DEFAULT '',
                    tipo_servicio VARCHAR(20) NOT NULL,
                    tipo_unidad VARCHAR(50),
                    peso_kg FLOAT,
                    kg_por_entrega FLOAT,
                    m3_por_entrega FLOAT,
                    temperatura VARCHAR(50),
                    custodia VARCHAR(30),
                    comentarios TEXT,
                    created_at TIMESTAMP DEFAULT NOW()
                )""",
                "ALTER TABLE rutas_transporte ADD COLUMN IF NOT EXISTS origen_estado VARCHAR(100) DEFAULT ''",
                "ALTER TABLE rutas_transporte ADD COLUMN IF NOT EXISTS origen_ciudad VARCHAR(100) DEFAULT ''",
                "ALTER TABLE rutas_transporte ADD COLUMN IF NOT EXISTS destino_estado VARCHAR(100) DEFAULT ''",
                "ALTER TABLE rutas_transporte ADD COLUMN IF NOT EXISTS destino_ciudad VARCHAR(100) DEFAULT ''",
                "ALTER TABLE rutas_transporte ADD COLUMN IF NOT EXISTS peso_kg FLOAT",
                "ALTER TABLE rutas_transporte ADD COLUMN IF NOT EXISTS kg_por_entrega FLOAT",
                "ALTER TABLE rutas_transporte ADD COLUMN IF NOT EXISTS m3_por_entrega FLOAT",
            ]
            for sql in migraciones:
                try:
                    conn.execute(text(sql))
                except Exception:
                    pass
            conn.commit()
        db.create_all()
        init_db()
        _db_initialized = True
    except Exception as e:
        print(f"Lazy DB init falló (reintentará en próximo request): {e}")

login_manager.login_message = 'Inicia sesión para continuar.'
login_manager.login_message_category = 'warning'

ROLES_LABEL = {
    'administrador':    'Administrador',
    'lider_comercial':  'Líder Comercial',
    'lider_soluciones': 'Líder de Soluciones',
    'aux_comercial':    'Aux Comercial',
    'comercial':        'Comercial',
    'ingeniero':        'Ingeniero',
    'aux_ingenieria':   'Aux Ingeniería',
}

TEMAS_DEFAULT = ['Reingeniería','Transporte','Almacenaje','VAS','Transporte y almacenaje']

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


# ── Decoradores ──────────────────────────────────────────────────────────────
def rol_requerido(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if current_user.rol not in roles:
                flash('No tienes permiso para realizar esta acción.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated
    return decorator


# ── Helpers ──────────────────────────────────────────────────────────────────
def es_admin():
    return current_user.rol == 'administrador'

def es_lider():
    return current_user.rol in ('administrador','lider_comercial','lider_soluciones','aux_comercial')

def es_comercial():
    return current_user.rol in ('administrador','comercial','lider_comercial','aux_comercial')

def es_soluciones():
    return current_user.rol in ('administrador','ingeniero','aux_ingenieria','lider_soluciones')

def puede_asignar_ingeniero():
    return current_user.rol in ('administrador','lider_soluciones')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_upload_path(folio, tipo):
    path = os.path.join(UPLOAD_BASE, folio, tipo)
    os.makedirs(path, exist_ok=True)
    return path

def generar_folio():
    anio = cdmx_now().year
    ultima = (Solicitud.query
              .filter(Solicitud.folio.like(f'SOL-{anio}-%'))
              .order_by(Solicitud.id.desc()).first())
    num = int(ultima.folio.split('-')[-1]) + 1 if ultima else 1
    return f'SOL-{anio}-{num:04d}'

def registrar_bitacora(solicitud_id, accion, usuario_id=None):
    db.session.add(Bitacora(
        solicitud_id=solicitud_id,
        usuario_id=usuario_id or (current_user.id if current_user.is_authenticated else None),
        accion=accion
    ))

def get_temas():
    return TemasSolicitud.query.filter_by(activo=True).order_by(TemasSolicitud.orden, TemasSolicitud.nombre).all()

def asignar_ingeniero_automatico():
    """Asigna al ingeniero con menor número de solicitudes activas."""
    ingenieros = User.query.filter(User.rol.in_(['ingeniero','aux_ingenieria']), User.activo==True).all()
    if not ingenieros:
        return None
    cargas = []
    for ing in ingenieros:
        activas = Solicitud.query.filter_by(
            responsable_id=ing.id
        ).filter(Solicitud.estatus != 'Cerrada').count()
        cargas.append((ing.id, activas))
    # Ordenar por menor carga, en caso de empate tomar el primero
    cargas.sort(key=lambda x: x[1])
    return cargas[0][0]


# ── Auth ─────────────────────────────────────────────────────────────────────
@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username, activo=True).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user, remember=True)
            return redirect(url_for('dashboard'))
        flash('Usuario o contraseña incorrectos.', 'danger')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


# ── Recuperar contraseña ─────────────────────────────────────────────────────
@app.route('/recuperar', methods=['GET', 'POST'])
def recuperar_password():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        nueva    = request.form.get('nueva_password', '')
        confirm  = request.form.get('confirm_password', '')
        user = User.query.filter_by(username=username, activo=True).first()
        if not user:
            flash('Usuario no encontrado.', 'danger')
        elif len(nueva) < 6:
            flash('La contraseña debe tener al menos 6 caracteres.', 'danger')
        elif nueva != confirm:
            flash('Las contraseñas no coinciden.', 'danger')
        else:
            user.password_hash = generate_password_hash(nueva)
            db.session.commit()
            flash('Contraseña actualizada. Ya puedes iniciar sesión.', 'success')
            return redirect(url_for('login'))
    return render_template('recuperar_password.html')


# ── Dashboard ────────────────────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    q = Solicitud.query
    # Todos los roles ven TODAS las solicitudes en el dashboard

    # Filtros dashboard
    f_estatus_list   = request.args.getlist('f_estatus')
    f_estatus        = f_estatus_list[0] if len(f_estatus_list)==1 else ''
    f_comercial_list = request.args.getlist('f_comercial')
    f_comercial      = f_comercial_list[0] if len(f_comercial_list)==1 else ''
    f_ingeniero_list = request.args.getlist('f_ingeniero')
    f_tema           = request.args.get('f_tema','').strip()

    if f_estatus_list:   q = q.filter(Solicitud.estatus.in_(f_estatus_list))
    if f_comercial_list: q = q.filter(Solicitud.hunter_id.in_([int(i) for i in f_comercial_list]))
    if f_ingeniero_list: q = q.filter(Solicitud.responsable_id.in_([int(i) for i in f_ingeniero_list]))
    if f_tema:           q = q.filter(Solicitud.tema == f_tema)

    todas = q.all()
    # Propuesta Enviada se trata como cerrada
    ESTATUS_CERRADO = ('Cerrada', 'Enviada')
    abiertas    = [s for s in todas if s.estatus not in ESTATUS_CERRADO]
    cerradas    = [s for s in todas if s.estatus in ESTATUS_CERRADO]
    vencidas    = [s for s in todas if s.dias_sin_movimiento() > 15 and s.estatus not in ESTATUS_CERRADO]
    monto_total = sum(s.monto_oportunidad or 0 for s in todas)

    tiempos = [s.dias_desde_captura() for s in cerradas]
    promedio_atencion = round(sum(tiempos)/len(tiempos), 1) if tiempos else 0

    # KPI Cumplimiento — solicitudes que llegaron a Info Completa antes de fecha compromiso
    con_compromiso = [s for s in todas if s.fecha_info_completa and s.fecha_compromiso]
    cumplidas = [s for s in con_compromiso if s.fecha_info_completa.date() <= s.fecha_compromiso]
    pct_cumplimiento = round(len(cumplidas)/len(con_compromiso)*100) if con_compromiso else None

    # KPI Calidad de info comercial — promedio de checkboxes completados
    def _calidad(s):
        checks = [s.historial_surtido, s.inventario, s.maestro_productos,
                  s.historial_recepcion, s.cuestionario_logistico]
        return round(sum(1 for c in checks if c) / len(checks) * 100)
    con_info = [s for s in todas if s.estatus not in ('Asignada',)]
    calidad_promedio = round(sum(_calidad(s) for s in con_info)/len(con_info)) if con_info else None

    # Tiempo promedio por estatus (cuello de botella)
    ESTATUS_FLOW = ['Asignada','En Análisis','Pendiente de Información','Información Completa',
                    'En Proceso','Revisada por Área Comercial','Pendiente de Liberación DG',
                    'Liberada','Enviada','Cerrada']

    estatus_list = ['Asignada','En Análisis','Pendiente de Información','Información Completa',
                    'En Proceso','Revisada por Área Comercial','Pendiente de Liberación DG',
                    'Liberada','Enviada','Cerrada']

    # Donut y gráficas respetan los filtros activos
    por_estatus = {e: len([s for s in todas if s.estatus == e]) for e in estatus_list}

    por_comercial_monto = {}
    por_comercial_count = {}
    por_ingeniero_monto = {}
    por_ingeniero_data  = {}
    if es_lider():
        from collections import defaultdict
        com_montos = defaultdict(float)
        com_counts = defaultdict(int)
        ing_data   = defaultdict(lambda: {'count':0,'monto_total':0,'solicitudes':[]})
        for s in todas:
            if s.estatus not in ESTATUS_CERRADO:
                monto = s.monto_oportunidad or 0
                if s.hunter:
                    com_montos[s.hunter.nombre] += monto
                    com_counts[s.hunter.nombre] += 1
                if s.responsable and s.responsable.rol in ('ingeniero','aux_ingenieria'):
                    key = s.responsable.nombre
                    ing_data[key]['count'] += 1
                    ing_data[key]['monto_total'] += monto
                    ing_data[key]['solicitudes'].append({
                        'folio': s.folio, 'cliente': s.cliente, 'monto': monto
                    })
        por_comercial_monto = {k: round(v) for k, v in
            sorted(com_montos.items(), key=lambda x: x[1], reverse=True) if v}
        por_comercial_count = {k: com_counts[k] for k in por_comercial_monto}
        por_ingeniero_monto = {k: round(d['monto_total']) for k, d in
            sorted(ing_data.items(), key=lambda x: x[1]['count'], reverse=True) if d['count']}
        por_ingeniero_data  = dict(sorted(ing_data.items(),
            key=lambda x: x[1]['count'], reverse=True))

    ultimas = sorted(todas, key=lambda s: s.ultima_actualizacion or s.fecha_captura, reverse=True)[:5]
    comerciales_list = User.query.filter(User.rol.in_(['comercial','aux_comercial','lider_comercial'])).all()
    temas_list = get_temas()

    return render_template('dashboard.html',
        total=len(todas), abiertas=len(abiertas), cerradas=len(cerradas),
        vencidas=len(vencidas), monto_total=monto_total,
        promedio_atencion=promedio_atencion,
        por_comercial_monto=por_comercial_monto,
        por_ingeniero_monto=por_ingeniero_monto,
        por_estatus=por_estatus, ultimas=ultimas,
        ROLES_LABEL=ROLES_LABEL,
        estatus_list=estatus_list,
        comerciales_list=comerciales_list,
        temas_list=temas_list,
        pct_cumplimiento=pct_cumplimiento,
        calidad_promedio=calidad_promedio,
        por_comercial_count=por_comercial_count,
        por_ingeniero_data=por_ingeniero_data,
        f_estatus=f_estatus, f_estatus_list=f_estatus_list,
        f_comercial=f_comercial, f_comercial_list=f_comercial_list,
        f_ingeniero_list=f_ingeniero_list,
        f_tema=f_tema,
        ingenieros_list=User.query.filter(User.rol.in_(['ingeniero','aux_ingenieria']),User.activo==True).order_by(User.nombre).all())


# ── Solicitudes ──────────────────────────────────────────────────────────────
@app.route('/solicitudes')
@login_required
def solicitudes():
    q = Solicitud.query
    # administrador, lider_soluciones e ingenieros ven TODAS sin restricción
    if current_user.rol in ('lider_comercial','aux_comercial'):
        ids = [u.id for u in User.query.filter(User.rol.in_(['comercial','aux_comercial'])).all()]
        ids.append(current_user.id)
        q = q.filter(Solicitud.hunter_id.in_(ids))

    folio       = request.args.get('folio','').strip()
    cliente     = request.args.get('cliente','').strip()
    estatus_list_f = request.args.getlist('estatus')  # multi-select
    estatus        = estatus_list_f[0] if len(estatus_list_f)==1 else ''
    comercial_f = request.args.get('comercial','').strip()
    ingeniero_f = request.args.get('ingeniero','').strip()
    tema_f      = request.args.get('tema','').strip()

    comercial_ids_f  = request.args.getlist('f_comercial_id')
    ingeniero_ids_f  = request.args.getlist('f_ingeniero_id')

    if folio:           q = q.filter(Solicitud.folio.ilike(f'%{folio}%'))
    if cliente:         q = q.filter(Solicitud.cliente.ilike(f'%{cliente}%'))
    if estatus_list_f:  q = q.filter(Solicitud.estatus.in_(estatus_list_f))
    if tema_f:          q = q.filter(Solicitud.tema == tema_f)
    if comercial_ids_f: q = q.filter(Solicitud.hunter_id.in_([int(i) for i in comercial_ids_f]))
    if ingeniero_ids_f: q = q.filter(Solicitud.responsable_id.in_([int(i) for i in ingeniero_ids_f]))

    lista      = q.order_by(Solicitud.fecha_captura.desc()).all()
    ingenieros = User.query.filter(User.rol.in_(['ingeniero','aux_ingenieria']), User.activo==True).all()
    estatus_list = ['Asignada','En Análisis','Pendiente de Información','Información Completa',
                    'En Proceso','Revisada por Área Comercial','Pendiente de Liberación DG',
                    'Liberada','Enviada','Cerrada']
    comerciales_sol = User.query.filter(
        User.rol.in_(['comercial','aux_comercial','lider_comercial']),
        User.activo==True).order_by(User.nombre).all()
    return render_template('solicitudes.html', solicitudes=lista,
                           ingenieros=ingenieros, estatus_list=estatus_list,
                           temas_list=get_temas(),
                           comerciales_list=comerciales_sol)


@app.route('/solicitudes/nueva', methods=['GET', 'POST'])
@login_required
@rol_requerido('comercial','lider_comercial','aux_comercial','administrador')
def nueva_solicitud():
    ingenieros  = User.query.filter(User.rol.in_(['ingeniero','aux_ingenieria']), User.activo==True).all()
    comerciales = User.query.filter(User.rol.in_(['comercial','aux_comercial','lider_comercial']), User.activo==True).all()
    temas       = get_temas()

    if request.method == 'POST':
        f = request.form
        try:
            fecha_sol = datetime.strptime(f['fecha_solicitud'], '%Y-%m-%d').date()
        except (ValueError, KeyError):
            flash('Fecha de solicitud inválida.', 'danger')
            return render_template('nueva_solicitud.html', ingenieros=ingenieros,
                                   comerciales=comerciales, temas=temas)

        # Fecha no puede ser anterior a hoy
        if fecha_sol < date.today():
            flash('La fecha de solicitud no puede ser anterior a hoy.', 'danger')
            return render_template('nueva_solicitud.html', ingenieros=ingenieros,
                                   comerciales=comerciales, temas=temas)

        # Subtipo (solo si tema incluye transporte/almacenaje)
        tema_val = f['tema'].strip().lower()
        subtipo  = f.get('subtipo','').strip() or None
        if not any(p in tema_val for p in ['transporte','almacenaje']):
            subtipo = None

        # Auto-asignación por menor carga de trabajo
        responsable_id = asignar_ingeniero_automatico()

        prio_sug = None  # Prioridad asignada por Líder de Soluciones

        origen_id = f.get('solicitud_origen_id','').strip()
        origen_id = int(origen_id) if origen_id else None

        sol = Solicitud(
            folio=generar_folio(),
            hunter_id=int(f.get('hunter_id', current_user.id)),
            responsable_id=responsable_id,
            solicitud_origen_id=origen_id,
            fecha_solicitud=fecha_sol,
            cliente=f['cliente'].strip(),
            tema=f['tema'].strip(),
            subtipo=subtipo,
            comentarios_comerciales=f.get('comentarios_comerciales','').strip(),
            monto_oportunidad=float(f['monto_oportunidad'].replace(',','')) if f.get('monto_oportunidad') else None,
            prioridad_sugerida=None,
            prioridad_comercial=None,
            prioridad=None,
            prioridad_estado='pendiente',
            estatus='Asignada' if responsable_id else 'Asignada',
        )
        db.session.add(sol)
        db.session.flush()
        registrar_bitacora(sol.id, f'{current_user.nombre} creó la solicitud.')
        # Registrar ingeniero principal en tabla intermedia
        if responsable_id:
            db.session.add(SolicitudIngeniero(
                solicitud_id=sol.id,
                ingeniero_id=responsable_id,
                es_principal=True
            ))
        # Si se liga a solicitud cerrada, copiar rutas
        if origen_id:
            sol_origen = Solicitud.query.get(origen_id)
            if sol_origen:
                registrar_bitacora(sol.id, f'Retrabajo ligado a {sol_origen.folio}.')

        # Rutas de transporte (solo si el tema incluye transporte)
        tema_lower = sol.tema.lower()
        es_transporte = 'transporte' in tema_lower
        if es_transporte:
            origenes      = request.form.getlist('ruta_origen[]')
            destinos      = request.form.getlist('ruta_destino[]')
            servicios     = request.form.getlist('ruta_tipo_servicio[]')
            unidades      = request.form.getlist('ruta_tipo_unidad[]')
            pesos         = request.form.getlist('ruta_peso[]')
            temperaturas  = request.form.getlist('ruta_temperatura[]')
            custodias     = request.form.getlist('ruta_custodia[]')
            comentarios_r = request.form.getlist('ruta_comentarios[]')

            o_estados = request.form.getlist('ruta_origen_estado[]')
            o_ciudades = request.form.getlist('ruta_origen_ciudad[]')
            d_estados  = request.form.getlist('ruta_destino_estado[]')
            d_ciudades = request.form.getlist('ruta_destino_ciudad[]')
            pesos_kg   = request.form.getlist('ruta_peso_kg[]')
            kg_ent     = request.form.getlist('ruta_kg_por_entrega[]')
            m3_ent     = request.form.getlist('ruta_m3_por_entrega[]')

            for i in range(len(o_estados)):
                oe = o_estados[i].strip() if i < len(o_estados) else ''
                oc = o_ciudades[i].strip() if i < len(o_ciudades) else ''
                de = d_estados[i].strip() if i < len(d_estados) else ''
                dc = d_ciudades[i].strip() if i < len(d_ciudades) else ''
                if oe and oc and de and dc:
                    svc = servicios[i] if i < len(servicios) else 'FTL'
                    def _f(lst, idx):
                        try: v = lst[idx].strip(); return float(v) if v else None
                        except: return None
                    ruta = RutaTransporte(
                        solicitud_id=sol.id, orden=i+1,
                        origen_estado=oe, origen_ciudad=oc,
                        destino_estado=de, destino_ciudad=dc,
                        tipo_servicio=svc,
                        tipo_unidad=unidades[i].strip() if svc != 'LTL' and i < len(unidades) else None,
                        peso_kg=_f(pesos_kg, i) if svc != 'LTL' else None,
                        kg_por_entrega=_f(kg_ent, i) if svc == 'LTL' else None,
                        m3_por_entrega=_f(m3_ent, i) if svc == 'LTL' else None,
                        temperatura=temperaturas[i].strip() if i < len(temperaturas) else None,
                        custodia=custodias[i].strip() if i < len(custodias) else None,
                        comentarios=comentarios_r[i].strip() if i < len(comentarios_r) else None,
                    )
                    db.session.add(ruta)
            if es_transporte and origenes:
                registrar_bitacora(sol.id, f'{current_user.nombre} registró {len([o for o in origenes if o.strip()])} ruta(s) de transporte.')

        # Archivos adjuntos al crear (comercial)
        archivos = request.files.getlist('archivos_iniciales')
        for archivo in archivos:
            if archivo and archivo.filename and allowed_file(archivo.filename):
                nombre_original = secure_filename(archivo.filename)
                ext = nombre_original.rsplit('.', 1)[1].lower()
                nombre_guardado = f"{uuid.uuid4().hex}.{ext}"
                ruta = get_upload_path(sol.folio, 'comercial')
                archivo.save(os.path.join(ruta, nombre_guardado))
                doc = Documento(
                    solicitud_id=sol.id, nombre_original=nombre_original,
                    nombre_guardado=nombre_guardado, tipo_documento='comercial',
                    usuario_id=current_user.id, version=1, activo=True,
                )
                db.session.add(doc)
                registrar_bitacora(sol.id, f'{current_user.nombre} adjuntó "{nombre_original}" al crear la solicitud.')

        db.session.commit()
        flash(f'Solicitud {sol.folio} creada exitosamente.', 'success')
        return redirect(url_for('detalle_solicitud', folio=sol.folio))

    return render_template('nueva_solicitud.html', ingenieros=ingenieros,
                           comerciales=comerciales, temas=temas)


@app.route('/solicitudes/<folio>')
@login_required
def detalle_solicitud(folio):
    sol = Solicitud.query.filter_by(folio=folio).first_or_404()
    # administrador y lider_soluciones acceden a cualquier solicitud sin filtro
    if current_user.rol in ('lider_comercial', 'aux_comercial'):
        ids = [u.id for u in User.query.filter(User.rol.in_(['comercial','aux_comercial'])).all()]
        ids.append(current_user.id)
        if sol.hunter_id not in ids:
            flash('No tienes acceso a esta solicitud.', 'danger')
            return redirect(url_for('solicitudes'))
    ingenieros   = User.query.filter(User.rol.in_(['ingeniero','aux_ingenieria']), User.activo==True).all()
    estatus_list = ['Asignada','En Análisis','Pendiente de Información','Información Completa',
                    'En Proceso','Revisada por Área Comercial','Pendiente de Liberación DG',
                    'Liberada','Enviada','Cerrada']
    return render_template('detalle_solicitud.html', sol=sol,
                           ingenieros=ingenieros, estatus_list=estatus_list)


@app.route('/solicitudes/<folio>/actualizar', methods=['POST'])
@login_required
@rol_requerido('ingeniero','aux_ingenieria','aux_ingenieria','lider_soluciones','administrador')
def actualizar_solicitud(folio):
    sol = Solicitud.query.filter_by(folio=folio).first_or_404()
    if current_user.rol in ('ingeniero','aux_ingenieria') and sol.responsable_id != current_user.id and not es_admin():
        flash('No tienes permiso para actualizar esta solicitud.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio))

    f = request.form
    cambios = []
    checkboxes = {
        'historial_surtido':'Historial de Surtido','inventario':'Inventario',
        'maestro_productos':'Maestro de Productos','historial_recepcion':'Historial de Recepción',
        'cuestionario_logistico':'Cuestionario Logístico',
    }
    for campo, label in checkboxes.items():
        nuevo = campo in f
        if getattr(sol, campo) != nuevo:
            setattr(sol, campo, nuevo)
            cambios.append(f'{current_user.nombre} {"marcó" if nuevo else "desmarcó"} {label}.')

    if f.get('fecha_compromiso'):
        try:
            nf = datetime.strptime(f['fecha_compromiso'], '%Y-%m-%d').date()
            if sol.fecha_compromiso != nf:
                sol.fecha_compromiso = nf
                cambios.append(f'{current_user.nombre} actualizó Fecha Compromiso a {nf.strftime("%d/%m/%Y")}.')
        except ValueError:
            pass

    if f.get('estatus'):
        nuevo_est = f['estatus']
        if sol.estatus != nuevo_est and nuevo_est != 'Cerrada':
            sol.estatus = nuevo_est
            cambios.append(f'{current_user.nombre} cambió estatus a "{nuevo_est}".')

    sol.ultima_actualizacion = cdmx_now()
    sol.actualizar_estatus_automatico()
    for c in cambios:
        registrar_bitacora(sol.id, c)
    db.session.commit()
    flash('Solicitud actualizada.', 'success')
    return redirect(url_for('detalle_solicitud', folio=folio))


@app.route('/solicitudes/<folio>/envio', methods=['POST'])
@login_required
@rol_requerido('ingeniero','aux_ingenieria','aux_ingenieria','lider_soluciones','administrador')
def registrar_envio(folio):
    sol = Solicitud.query.filter_by(folio=folio).first_or_404()
    if not request.form.get('comentarios_envio','').strip():
        flash('Debes agregar un comentario del envío.', 'warning')
        return redirect(url_for('detalle_solicitud', folio=folio))
    sol.fecha_envio_cliente  = cdmx_now()
    sol.usuario_envio_id     = current_user.id
    sol.comentarios_envio    = request.form['comentarios_envio'].strip()
    sol.estatus              = 'Enviada'
    sol.ultima_actualizacion = cdmx_now()
    # Calcular días de análisis
    if sol.fecha_info_completa:
        sol.dias_analisis = (cdmx_now() - sol.fecha_info_completa).days
    registrar_bitacora(sol.id, f'{current_user.nombre} registró envío de propuesta al cliente.')
    db.session.commit()
    flash('Envío registrado exitosamente.', 'success')
    return redirect(url_for('detalle_solicitud', folio=folio))


@app.route('/solicitudes/<folio>/cerrar', methods=['POST'])
@login_required
@rol_requerido('lider_soluciones','administrador')
def cerrar_solicitud(folio):
    sol = Solicitud.query.filter_by(folio=folio).first_or_404()
    sol.estatus = 'Cerrada'; sol.fecha_cierre = cdmx_now()
    sol.usuario_cierre_id = current_user.id; sol.ultima_actualizacion = cdmx_now()
    registrar_bitacora(sol.id, f'{current_user.nombre} cerró la solicitud.')
    db.session.commit()
    flash('Solicitud cerrada.', 'success')
    return redirect(url_for('detalle_solicitud', folio=folio))


@app.route('/solicitudes/<folio>/reasignar', methods=['POST'])
@login_required
@rol_requerido('lider_soluciones','administrador')
def reasignar_solicitud(folio):
    sol = Solicitud.query.filter_by(folio=folio).first_or_404()
    nuevo_id = request.form.get('responsable_id')
    if nuevo_id:
        nuevo_resp = db.session.get(User, int(nuevo_id))
        if nuevo_resp:
            sol.responsable_id = nuevo_resp.id
            if sol.estatus == 'Capturada': sol.estatus = 'Asignada'
            sol.ultima_actualizacion = cdmx_now()
            registrar_bitacora(sol.id, f'{current_user.nombre} reasignó a {nuevo_resp.nombre}.')
            db.session.commit()
            flash(f'Reasignada a {nuevo_resp.nombre}.', 'success')
    return redirect(url_for('detalle_solicitud', folio=folio))


@app.route('/solicitudes/<folio>/comentario', methods=['POST'])
@login_required
def agregar_comentario(folio):
    sol  = Solicitud.query.filter_by(folio=folio).first_or_404()
    texto = request.form.get('texto','').strip()
    if not texto:
        flash('El comentario no puede estar vacío.', 'warning')
        return redirect(url_for('detalle_solicitud', folio=folio))
    db.session.add(Comentario(solicitud_id=sol.id, usuario_id=current_user.id, texto=texto))
    sol.ultima_actualizacion = cdmx_now()
    registrar_bitacora(sol.id, f'{current_user.nombre} agregó un comentario.')
    db.session.commit()
    flash('Comentario agregado.', 'success')
    return redirect(url_for('detalle_solicitud', folio=folio))


@app.route('/solicitudes/<folio>/eliminar', methods=['POST'])
@login_required
@rol_requerido('lider_comercial','administrador')
def eliminar_solicitud(folio):
    sol = Solicitud.query.filter_by(folio=folio).first_or_404()
    folio_g = sol.folio; cliente = sol.cliente
    Comentario.query.filter_by(solicitud_id=sol.id).delete()
    Bitacora.query.filter_by(solicitud_id=sol.id).delete()
    Documento.query.filter_by(solicitud_id=sol.id).delete()
    db.session.delete(sol)
    db.session.commit()
    flash(f'Solicitud {folio_g} ({cliente}) eliminada.', 'success')
    return redirect(url_for('solicitudes'))


# ── Admin: Temas ─────────────────────────────────────────────────────────────
@app.route('/admin/temas')
@login_required
@rol_requerido('administrador')
def admin_temas():
    temas = TemasSolicitud.query.order_by(TemasSolicitud.orden, TemasSolicitud.nombre).all()
    return render_template('admin_temas.html', temas=temas)


@app.route('/admin/temas/nuevo', methods=['POST'])
@login_required
@rol_requerido('administrador')
def admin_nuevo_tema():
    nombre = request.form.get('nombre','').strip()
    if not nombre:
        flash('El nombre es obligatorio.', 'danger')
    elif TemasSolicitud.query.filter_by(nombre=nombre).first():
        flash(f'"{nombre}" ya existe.', 'danger')
    else:
        db.session.add(TemasSolicitud(nombre=nombre))
        db.session.commit()
        flash(f'Tema "{nombre}" agregado.', 'success')
    return redirect(url_for('admin_temas'))


@app.route('/admin/temas/<int:tid>/toggle', methods=['POST'])
@login_required
@rol_requerido('administrador')
def admin_toggle_tema(tid):
    tema = db.session.get(TemasSolicitud, tid)
    if tema:
        tema.activo = not tema.activo
        db.session.commit()
        flash(f'Tema "{tema.nombre}" {"activado" if tema.activo else "desactivado"}.', 'success')
    return redirect(url_for('admin_temas'))


@app.route('/admin/temas/<int:tid>/eliminar', methods=['POST'])
@login_required
@rol_requerido('administrador')
def admin_eliminar_tema(tid):
    tema = db.session.get(TemasSolicitud, tid)
    if tema:
        db.session.delete(tema)
        db.session.commit()
        flash('Tema eliminado.', 'success')
    return redirect(url_for('admin_temas'))


# ── Admin: Usuarios ───────────────────────────────────────────────────────────
@app.route('/admin/usuarios')
@login_required
@rol_requerido('lider_comercial','lider_soluciones','aux_comercial','administrador')
def admin_usuarios():
    if es_admin():
        roles_visibles = list(ROLES_LABEL.keys())
    elif current_user.rol in ('lider_comercial','aux_comercial'):
        roles_visibles = ['comercial','aux_comercial','lider_comercial']
    else:
        roles_visibles = ['ingeniero','aux_ingenieria','lider_soluciones']
    usuarios = User.query.filter(User.rol.in_(roles_visibles)).order_by(User.nombre).all()
    return render_template('admin_usuarios.html', usuarios=usuarios,
                           roles_visibles=roles_visibles, ROLES_LABEL=ROLES_LABEL)


@app.route('/admin/usuarios/nuevo', methods=['GET','POST'])
@login_required
@rol_requerido('lider_comercial','lider_soluciones','aux_comercial','administrador')
def admin_nuevo_usuario():
    if es_admin():
        roles_permitidos = list(ROLES_LABEL.keys())
    elif current_user.rol in ('lider_comercial','aux_comercial'):
        roles_permitidos = ['comercial','aux_comercial']
    else:
        roles_permitidos = ['ingeniero']

    if request.method == 'POST':
        f = request.form
        username = f.get('username','').strip().lower()
        nombre   = f.get('nombre','').strip()
        rol      = f.get('rol','')
        password = f.get('password','')
        confirm  = f.get('confirm','')

        if not all([username, nombre, rol, password]):
            flash('Todos los campos son obligatorios.', 'danger')
        elif rol not in roles_permitidos:
            flash('Rol no permitido.', 'danger')
        elif password != confirm:
            flash('Las contraseñas no coinciden.', 'danger')
        elif len(password) < 6:
            flash('Mínimo 6 caracteres.', 'danger')
        elif User.query.filter_by(username=username).first():
            flash(f'El usuario "{username}" ya existe.', 'danger')
        else:
            db.session.add(User(username=username, nombre=nombre, rol=rol,
                                password_hash=generate_password_hash(password)))
            db.session.commit()
            flash(f'Usuario {nombre} creado.', 'success')
            return redirect(url_for('admin_usuarios'))

    return render_template('admin_form_usuario.html', usuario=None,
                           roles_permitidos=roles_permitidos,
                           ROLES_LABEL=ROLES_LABEL, accion='nuevo')


@app.route('/admin/usuarios/<int:uid>/editar', methods=['GET','POST'])
@login_required
@rol_requerido('lider_comercial','lider_soluciones','aux_comercial','administrador')
def admin_editar_usuario(uid):
    usuario = db.session.get(User, uid)
    if not usuario:
        flash('Usuario no encontrado.', 'danger')
        return redirect(url_for('admin_usuarios'))

    if es_admin():
        roles_permitidos = list(ROLES_LABEL.keys())
    elif current_user.rol in ('lider_comercial','aux_comercial'):
        roles_permitidos = ['comercial','aux_comercial']
    else:
        roles_permitidos = ['ingeniero']

    if not es_admin() and usuario.rol not in roles_permitidos and usuario.id != current_user.id:
        flash('No puedes editar este usuario.', 'danger')
        return redirect(url_for('admin_usuarios'))

    if request.method == 'POST':
        f        = request.form
        nombre   = f.get('nombre','').strip()
        username = f.get('username','').strip().lower()
        password = f.get('password','').strip()
        confirm  = f.get('confirm','').strip()
        activo   = 'activo' in f
        existe   = User.query.filter_by(username=username).first()

        if not nombre:
            flash('El nombre es obligatorio.', 'danger')
        elif not username:
            flash('El usuario es obligatorio.', 'danger')
        elif existe and existe.id != usuario.id:
            flash(f'El usuario "{username}" ya está en uso.', 'danger')
        elif password and password != confirm:
            flash('Las contraseñas no coinciden.', 'danger')
        elif password and len(password) < 6:
            flash('Mínimo 6 caracteres.', 'danger')
        else:
            usuario.nombre = nombre; usuario.username = username; usuario.activo = activo
            if password:
                usuario.password_hash = generate_password_hash(password)
            db.session.commit()
            flash(f'Usuario {nombre} actualizado.', 'success')
            return redirect(url_for('admin_usuarios'))

    return render_template('admin_form_usuario.html', usuario=usuario,
                           roles_permitidos=roles_permitidos,
                           ROLES_LABEL=ROLES_LABEL, accion='editar')


@app.route('/admin/usuarios/<int:uid>/toggle', methods=['POST'])
@login_required
@rol_requerido('lider_comercial','lider_soluciones','aux_comercial','administrador')
def admin_toggle_usuario(uid):
    usuario = db.session.get(User, uid)
    if not usuario or usuario.id == current_user.id:
        flash('Operación no permitida.', 'danger')
        return redirect(url_for('admin_usuarios'))
    if not es_admin():
        if current_user.rol in ('lider_comercial','aux_comercial') and usuario.rol not in ['comercial','aux_comercial']:
            flash('No puedes modificar este usuario.', 'danger')
            return redirect(url_for('admin_usuarios'))
        if current_user.rol == 'lider_soluciones' and usuario.rol not in ['ingeniero']:
            flash('No puedes modificar este usuario.', 'danger')
            return redirect(url_for('admin_usuarios'))
    usuario.activo = not usuario.activo
    db.session.commit()
    flash(f'Usuario {usuario.nombre} {"activado" if usuario.activo else "desactivado"}.', 'success')
    return redirect(url_for('admin_usuarios'))


# ── Documentos ────────────────────────────────────────────────────────────────
def puede_subir_doc(tipo):
    if current_user.rol == 'administrador':
        return True
    if tipo in ('comercial',) and current_user.rol in ('comercial','lider_comercial','aux_comercial'):
        return True
    if tipo == 'soluciones' and current_user.rol in ('ingeniero','aux_ingenieria','lider_soluciones','aux_comercial','lider_comercial'):
        return True
    if tipo == 'propuesta_final' and current_user.rol in ('comercial','lider_comercial'):
        return True
    if current_user.rol in ('lider_comercial','lider_soluciones','aux_comercial'):
        return True
    return False

LIMITE_PROPUESTA = 6  # Versiones máximas de propuesta final

def puede_subir_propuesta_final(sol):
    """Verifica si se puede subir una versión más de propuesta final."""
    activas = Documento.query.filter_by(
        solicitud_id=sol.id, tipo_documento='propuesta_final', activo=True).count()
    historico = Documento.query.filter_by(
        solicitud_id=sol.id, tipo_documento='propuesta_final').count()
    if historico < LIMITE_PROPUESTA:
        return True, None
    # Over limit — only lider_comercial can authorize
    if current_user.rol in ('lider_comercial','administrador'):
        return True, 'autorizado_lider'
    return False, 'limite'


@app.route('/solicitudes/<folio>/documentos/subir', methods=['POST'])
@login_required
def subir_documento(folio):
    sol  = Solicitud.query.filter_by(folio=folio).first_or_404()
    tipo = request.form.get('tipo_documento','')
    if tipo not in ('comercial','soluciones'):
        flash('Tipo inválido.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')
    if not puede_subir_doc(tipo):
        flash('Sin permiso.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')
    archivo = request.files.get('archivo')
    if not archivo or archivo.filename == '':
        flash('No seleccionaste archivo.', 'warning')
        return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')
    if not allowed_file(archivo.filename):
        flash('Formato no permitido.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')
    nombre_original = secure_filename(archivo.filename)
    ext = nombre_original.rsplit('.', 1)[1].lower()
    nombre_guardado = f"{uuid.uuid4().hex}.{ext}"
    archivo.save(os.path.join(get_upload_path(folio, tipo), nombre_guardado))
    db.session.add(Documento(solicitud_id=sol.id, nombre_original=nombre_original,
        nombre_guardado=nombre_guardado, tipo_documento=tipo,
        usuario_id=current_user.id, version=1, activo=True))
    registrar_bitacora(sol.id, f'{current_user.nombre} cargó "{nombre_original}" ({tipo}).')
    db.session.commit()
    flash(f'"{nombre_original}" subido.', 'success')
    return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')


@app.route('/solicitudes/<folio>/documentos/<int:doc_id>/reemplazar', methods=['POST'])
@login_required
def reemplazar_documento(folio, doc_id):
    sol = Solicitud.query.filter_by(folio=folio).first_or_404()
    doc_ant = db.session.get(Documento, doc_id)
    if not doc_ant or not doc_ant.activo:
        flash('Documento no encontrado.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')
    if not puede_subir_doc(doc_ant.tipo_documento):
        flash('Sin permiso.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')
    archivo = request.files.get('archivo')
    if not archivo or archivo.filename == '' or not allowed_file(archivo.filename):
        flash('Archivo inválido.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')
    doc_ant.activo = False
    nombre_original = secure_filename(archivo.filename)
    ext = nombre_original.rsplit('.', 1)[1].lower()
    nombre_guardado = f"{uuid.uuid4().hex}.{ext}"
    archivo.save(os.path.join(get_upload_path(folio, doc_ant.tipo_documento), nombre_guardado))
    nueva = Documento(solicitud_id=sol.id, nombre_original=nombre_original,
        nombre_guardado=nombre_guardado, tipo_documento=doc_ant.tipo_documento,
        usuario_id=current_user.id, version=doc_ant.version+1,
        activo=True, documento_padre_id=doc_ant.id)
    db.session.add(nueva)
    registrar_bitacora(sol.id, f'{current_user.nombre} reemplazó "{doc_ant.nombre_original}" → "{nombre_original}" (v{nueva.version}).')
    db.session.commit()
    flash(f'Reemplazado. Nueva versión {nueva.version}.', 'success')
    return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')


@app.route('/solicitudes/<folio>/documentos/<int:doc_id>/eliminar', methods=['POST'])
@login_required
def eliminar_documento(folio, doc_id):
    sol = Solicitud.query.filter_by(folio=folio).first_or_404()
    doc = db.session.get(Documento, doc_id)
    if not doc or not doc.activo:
        flash('Documento no encontrado.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')
    if not puede_subir_doc(doc.tipo_documento):
        flash('Sin permiso.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')
    ruta_arch = os.path.join(get_upload_path(folio, doc.tipo_documento), doc.nombre_guardado)
    if os.path.exists(ruta_arch): os.remove(ruta_arch)
    nombre = doc.nombre_original; tipo = doc.tipo_documento
    db.session.delete(doc)
    registrar_bitacora(sol.id, f'{current_user.nombre} eliminó "{nombre}" ({tipo}).')
    db.session.commit()
    flash(f'"{nombre}" eliminado.', 'success')
    return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')


@app.route('/solicitudes/<folio>/documentos/<int:doc_id>/descargar')
@login_required
def descargar_documento(folio, doc_id):
    sol = Solicitud.query.filter_by(folio=folio).first_or_404()
    doc = db.session.get(Documento, doc_id)
    if not doc: abort(404)
    ruta = get_upload_path(folio, doc.tipo_documento)
    if not os.path.exists(os.path.join(ruta, doc.nombre_guardado)):
        flash('Archivo no encontrado.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')
    registrar_bitacora(sol.id, f'{current_user.nombre} descargó "{doc.nombre_original}" ({doc.tipo_documento}).')
    db.session.commit()
    return send_from_directory(ruta, doc.nombre_guardado,
                               as_attachment=True, download_name=doc.nombre_original)





@app.route('/solicitudes/<folio>/prioridad', methods=['POST'])
@login_required
@rol_requerido('comercial','lider_comercial','lider_soluciones','administrador','aux_comercial')
def actualizar_prioridad(folio):
    sol  = Solicitud.query.filter_by(folio=folio).first_or_404()
    paso = request.form.get('paso','')  # sugerir | confirmar_com | confirmar_sol
    valor = request.form.get('prioridad','').strip()

    try:
        num = int(valor) if valor else None
        if num is not None and num < 1:
            raise ValueError
    except ValueError:
        flash('La prioridad debe ser un número mayor a 0.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio))

    if paso == 'sugerir' and current_user.rol in ('comercial','aux_comercial','lider_comercial','administrador'):
        sol.prioridad_sugerida = num
        sol.prioridad_estado   = 'pendiente'
        registrar_bitacora(sol.id, f'{current_user.nombre} sugirió prioridad #{num}.')
        flash(f'Prioridad sugerida #{num}. Pendiente de confirmación por Líder Comercial.', 'success')

    elif paso == 'confirmar_com' and current_user.rol in ('lider_comercial','administrador'):
        sol.prioridad_comercial = num
        sol.prioridad_estado    = 'confirmada_com'
        registrar_bitacora(sol.id, f'{current_user.nombre} confirmó prioridad comercial #{num}.')
        flash(f'Prioridad comercial confirmada #{num}. Pendiente de confirmación por Líder de Soluciones.', 'success')

    elif paso == 'confirmar_sol' and current_user.rol in ('lider_soluciones','administrador'):
        sol.prioridad       = num
        sol.prioridad_estado = 'confirmada'
        registrar_bitacora(sol.id, f'{current_user.nombre} confirmó prioridad final #{num}.')
        flash(f'Prioridad final confirmada #{num}.', 'success')
    else:
        flash('No tienes permiso para este paso de prioridad.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio))

    sol.ultima_actualizacion = cdmx_now()
    db.session.commit()
    return redirect(url_for('detalle_solicitud', folio=folio))


@app.route('/solicitudes/<folio>/rutas/guardar', methods=['POST'])
@login_required
@rol_requerido('comercial','lider_comercial','aux_comercial','administrador')
def guardar_rutas(folio):
    sol = Solicitud.query.filter_by(folio=folio).first_or_404()

    # Verificar permiso — solo el comercial dueño o líderes
    if current_user.rol == 'comercial' and sol.hunter_id != current_user.id:
        flash('Sin permiso para editar esta solicitud.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio))

    # Eliminar rutas anteriores y reescribir
    RutaTransporte.query.filter_by(solicitud_id=sol.id).delete()

    o_estados  = request.form.getlist('ruta_origen_estado[]')
    o_ciudades = request.form.getlist('ruta_origen_ciudad[]')
    d_estados  = request.form.getlist('ruta_destino_estado[]')
    d_ciudades = request.form.getlist('ruta_destino_ciudad[]')
    servicios  = request.form.getlist('ruta_tipo_servicio[]')
    unidades   = request.form.getlist('ruta_tipo_unidad[]')
    pesos_kg   = request.form.getlist('ruta_peso_kg[]')
    kg_ent     = request.form.getlist('ruta_kg_por_entrega[]')
    m3_ent     = request.form.getlist('ruta_m3_por_entrega[]')
    temps      = request.form.getlist('ruta_temperatura[]')
    custodias  = request.form.getlist('ruta_custodia[]')
    coments    = request.form.getlist('ruta_comentarios[]')

    def _f(lst, i):
        try: v = lst[i].strip(); return float(v) if v else None
        except: return None

    count = 0
    for i in range(len(o_estados)):
        oe = o_estados[i].strip(); oc = o_ciudades[i].strip() if i < len(o_ciudades) else ''
        de = d_estados[i].strip() if i < len(d_estados) else ''
        dc = d_ciudades[i].strip() if i < len(d_ciudades) else ''
        if oe and oc and de and dc:
            svc = servicios[i] if i < len(servicios) else 'FTL'
            db.session.add(RutaTransporte(
                solicitud_id=sol.id, orden=count+1,
                origen_estado=oe, origen_ciudad=oc,
                destino_estado=de, destino_ciudad=dc,
                tipo_servicio=svc,
                tipo_unidad=unidades[i].strip() if svc != 'LTL' and i < len(unidades) else None,
                peso_kg=_f(pesos_kg, i) if svc != 'LTL' else None,
                kg_por_entrega=_f(kg_ent, i) if svc == 'LTL' else None,
                m3_por_entrega=_f(m3_ent, i) if svc == 'LTL' else None,
                temperatura=temps[i].strip() if i < len(temps) else None,
                custodia=custodias[i].strip() if i < len(custodias) else None,
                comentarios=coments[i].strip() if i < len(coments) else None,
            ))
            count += 1

    sol.ultima_actualizacion = cdmx_now()
    registrar_bitacora(sol.id, f'{current_user.nombre} actualizó las rutas de transporte ({count} rutas).')
    db.session.commit()
    flash(f'{count} ruta(s) de transporte guardadas correctamente.', 'success')
    return redirect(url_for('detalle_solicitud', folio=folio) + '#tab-transporte')


@app.route('/exportar/rutas-transporte')
@login_required
@rol_requerido('administrador','lider_comercial','lider_soluciones','aux_comercial')
def exportar_rutas():
    """Exporta todas las rutas de transporte a Excel."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Rutas de Transporte"

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="0F172A")
    alt_fill  = PatternFill("solid", fgColor="161B22")
    cell_border = Border(
        bottom=Side(style='thin', color='2A3040'),
        right=Side(style='thin', color='2A3040')
    )

    headers = [
        "Folio", "Cliente", "Tipo Solicitud", "Subtipo", "Comercial",
        "#", "Origen Estado", "Origen Ciudad", "Destino Estado", "Destino Ciudad",
        "Tipo Servicio", "Tipo Unidad", "Peso (kg)",
        "Kg/Entrega (LTL)", "M3/Entrega (LTL)",
        "Temperatura", "Custodia", "Comentarios"
    ]
    col_widths = [16,28,22,10,22,5,20,22,20,22,16,16,12,14,14,22,18,30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    # Only solicitudes with transporte tema
    folio_filter = request.args.get('folio','').strip()
    q = Solicitud.query.filter(Solicitud.tema.ilike('%transporte%'))
    if folio_filter:
        q = q.filter(Solicitud.folio == folio_filter)
    sols = q.order_by(Solicitud.id.desc()).all()

    row_num = 2
    for sol in sols:
        rutas = list(sol.rutas)
        if not rutas:
            continue
        fill = alt_fill if row_num % 2 == 0 else PatternFill("solid", fgColor="1A1F27")
        font = Font(color="E2E8F0", size=10)
        for r in rutas:
            row = [
                sol.folio, sol.cliente, sol.tema, sol.subtipo or '',
                sol.hunter.nombre if sol.hunter else '',
                r.orden,
                r.origen_estado, r.origen_ciudad,
                r.destino_estado, r.destino_ciudad,
                r.tipo_servicio, r.tipo_unidad or '',
                r.peso_kg or '',
                r.kg_por_entrega or '', r.m3_por_entrega or '',
                r.temperatura or '', r.custodia or '',
                r.comentarios or ''
            ]
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.font = font; cell.fill = fill
                cell.border = cell_border
                cell.alignment = Alignment(vertical="center")
            row_num += 1

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fecha = cdmx_now().strftime('%Y%m%d_%H%M')
    return Response(buf.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=RutasTransporte_{fecha}.xlsx'})


@app.route('/solicitudes/<folio>/ingenieros/agregar', methods=['POST'])
@login_required
@rol_requerido('lider_soluciones','administrador')
def agregar_ingeniero(folio):
    sol = Solicitud.query.filter_by(folio=folio).first_or_404()
    ing_id = request.form.get('ingeniero_id','').strip()
    if not ing_id:
        flash('Selecciona un ingeniero.', 'warning')
        return redirect(url_for('detalle_solicitud', folio=folio))
    ing_id = int(ing_id)
    # Check not already assigned
    ya = SolicitudIngeniero.query.filter_by(solicitud_id=sol.id, ingeniero_id=ing_id).first()
    if ya:
        flash('Ese ingeniero ya está asignado.', 'warning')
        return redirect(url_for('detalle_solicitud', folio=folio))
    db.session.add(SolicitudIngeniero(solicitud_id=sol.id, ingeniero_id=ing_id, es_principal=False))
    ing = db.session.get(User, ing_id)
    sol.ultima_actualizacion = cdmx_now()
    registrar_bitacora(sol.id, f'{current_user.nombre} agregó a {ing.nombre} como ingeniero colaborador.')
    db.session.commit()
    flash(f'{ing.nombre} agregado como colaborador.', 'success')
    return redirect(url_for('detalle_solicitud', folio=folio))


@app.route('/solicitudes/<folio>/ingenieros/<int:si_id>/quitar', methods=['POST'])
@login_required
@rol_requerido('lider_soluciones','administrador')
def quitar_ingeniero(folio, si_id):
    asig = db.session.get(SolicitudIngeniero, si_id)
    if not asig or asig.es_principal:
        flash('No puedes quitar al ingeniero principal desde aquí. Usa Reasignar.', 'warning')
        return redirect(url_for('detalle_solicitud', folio=folio))
    ing = db.session.get(User, asig.ingeniero_id)
    sol = Solicitud.query.filter_by(folio=folio).first_or_404()
    db.session.delete(asig)
    sol.ultima_actualizacion = cdmx_now()
    registrar_bitacora(sol.id, f'{current_user.nombre} quitó a {ing.nombre} de la solicitud.')
    db.session.commit()
    flash(f'{ing.nombre} quitado de la solicitud.', 'success')
    return redirect(url_for('detalle_solicitud', folio=folio))


@app.route('/api/solicitudes-cerradas')
@login_required
def api_solicitudes_cerradas():
    """API para buscar solicitudes cerradas — usado en formulario de retrabajo."""
    q = request.args.get('q','').strip()
    sols = Solicitud.query.filter(
        Solicitud.estatus == 'Cerrada',
        db.or_(
            Solicitud.folio.ilike(f'%{q}%'),
            Solicitud.cliente.ilike(f'%{q}%')
        )
    ).limit(10).all()
    return jsonify([{
        'id': s.id, 'folio': s.folio, 'cliente': s.cliente,
        'tema': s.tema, 'subtipo': s.subtipo or '',
        'monto': s.monto_oportunidad or 0,
        'comentarios': s.comentarios_comerciales or ''
    } for s in sols])


@app.route('/solicitudes/<folio>/propuesta-final/subir', methods=['POST'])
@login_required
def subir_propuesta_final(folio):
    sol = Solicitud.query.filter_by(folio=folio).first_or_404()
    puede, razon = puede_subir_propuesta_final(sol)
    if not puede:
        flash(f'Límite de {LIMITE_PROPUESTA} versiones alcanzado. Solicita autorización al Líder Comercial.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')
    if not current_user.rol in ('comercial','lider_comercial','administrador'):
        flash('Sin permiso.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')

    archivo = request.files.get('archivo')
    if not archivo or archivo.filename == '' or not allowed_file(archivo.filename):
        flash('Archivo inválido.', 'danger')
        return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')

    nombre_original = secure_filename(archivo.filename)
    ext = nombre_original.rsplit('.', 1)[1].lower()
    nombre_guardado = f"{uuid.uuid4().hex}.{ext}"
    archivo.save(os.path.join(get_upload_path(folio, 'propuesta_final'), nombre_guardado))

    version = Documento.query.filter_by(
        solicitud_id=sol.id, tipo_documento='propuesta_final').count() + 1
    db.session.add(Documento(
        solicitud_id=sol.id, nombre_original=nombre_original,
        nombre_guardado=nombre_guardado, tipo_documento='propuesta_final',
        usuario_id=current_user.id, version=version, activo=True,
    ))
    msg_extra = ' [autorizado por líder]' if razon == 'autorizado_lider' else ''
    registrar_bitacora(sol.id, f'{current_user.nombre} subió propuesta final v{version}: "{nombre_original}"{msg_extra}.')
    db.session.commit()
    flash(f'Propuesta final v{version} subida correctamente.', 'success')
    return redirect(url_for('detalle_solicitud', folio=folio) + '#archivos')


@app.route('/admin/debug-startup')
@login_required
@rol_requerido('administrador')
def debug_startup():
    """Ruta temporal para diagnosticar errores de startup."""
    from sqlalchemy import text, inspect
    results = {}
    try:
        with db.engine.connect() as conn:
            # Check tables exist
            inspector = inspect(db.engine)
            tables = inspector.get_table_names()
            results['tables'] = tables

            # Check columns in solicitudes
            if 'solicitudes' in tables:
                cols = [c['name'] for c in inspector.get_columns('solicitudes')]
                results['solicitudes_cols'] = cols

            # Check solicitud_ingenieros
            if 'solicitud_ingenieros' in tables:
                cols_si = [c['name'] for c in inspector.get_columns('solicitud_ingenieros')]
                results['solicitud_ingenieros_cols'] = cols_si
            else:
                results['solicitud_ingenieros'] = 'TABLE MISSING'

    except Exception as e:
        results['error'] = str(e)

    return jsonify(results)

# ── Acciones en lote ──────────────────────────────────────────────────────────
@app.route('/solicitudes/accion-lote', methods=['POST'])
@login_required
@rol_requerido('lider_comercial','lider_soluciones','administrador')
def accion_lote():
    folios  = request.form.getlist('folios')
    accion  = request.form.get('accion')

    if not folios:
        flash('No seleccionaste ninguna solicitud.', 'warning')
        return redirect(url_for('solicitudes'))
    if accion not in ('eliminar','cerrar'):
        flash('Acción no válida.', 'danger')
        return redirect(url_for('solicitudes'))

    # Permiso: solo admin y lider_comercial pueden eliminar
    if accion == 'eliminar' and current_user.rol not in ('administrador','lider_comercial'):
        flash('No tienes permiso para eliminar solicitudes.', 'danger')
        return redirect(url_for('solicitudes'))

    # Permiso: solo admin y lider_soluciones pueden cerrar
    if accion == 'cerrar' and current_user.rol not in ('administrador','lider_soluciones'):
        flash('No tienes permiso para cerrar solicitudes.', 'danger')
        return redirect(url_for('solicitudes'))

    procesadas = 0
    omitidas   = 0

    for folio in folios:
        sol = Solicitud.query.filter_by(folio=folio).first()
        if not sol:
            continue

        if accion == 'eliminar':
            Comentario.query.filter_by(solicitud_id=sol.id).delete()
            Bitacora.query.filter_by(solicitud_id=sol.id).delete()
            Documento.query.filter_by(solicitud_id=sol.id).delete()
            db.session.delete(sol)
            procesadas += 1

        elif accion == 'cerrar':
            if sol.estatus == 'Cerrada':
                omitidas += 1
                continue
            sol.estatus           = 'Cerrada'
            sol.fecha_cierre      = cdmx_now()
            sol.usuario_cierre_id = current_user.id
            sol.ultima_actualizacion = cdmx_now()
            registrar_bitacora(sol.id, f'{current_user.nombre} cerró la solicitud (acción en lote).')
            procesadas += 1

    db.session.commit()

    msg = f'{procesadas} solicitud{"es" if procesadas != 1 else ""} '
    msg += 'eliminada' + ('s' if procesadas != 1 else '') if accion == 'eliminar'           else 'cerrada' + ('s' if procesadas != 1 else '')
    if omitidas:
        msg += f' ({omitidas} ya estaban cerradas, omitidas).'
    flash(msg + '.', 'success')
    return redirect(url_for('solicitudes'))

# ══════════════════════════════════════════════════════════════════════════════
# ── EXPORTACIONES ────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/exportar/solicitudes')
@login_required
@rol_requerido('administrador','lider_comercial','lider_soluciones','aux_comercial')
def exportar_solicitudes():
    q = Solicitud.query

    # Aplicar mismos filtros de rol que en la vista
    if current_user.rol in ('lider_comercial','aux_comercial'):
        ids = [u.id for u in User.query.filter(User.rol.in_(['comercial','aux_comercial'])).all()]
        ids.append(current_user.id)
        q = q.filter(Solicitud.hunter_id.in_(ids))
    elif current_user.rol == 'lider_soluciones':
        pass  # ve todas
    # administrador ve todas sin filtro

    # Filtros opcionales por querystring
    f_estatus   = request.args.get('f_estatus','').strip()
    f_comercial = request.args.get('f_comercial','').strip()
    f_tema      = request.args.get('f_tema','').strip()
    if f_estatus:   q = q.filter(Solicitud.estatus == f_estatus)
    if f_comercial: q = q.join(User, Solicitud.hunter_id == User.id).filter(User.id == int(f_comercial))
    if f_tema:      q = q.filter(Solicitud.tema == f_tema)

    solicitudes = q.order_by(Solicitud.fecha_captura.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes"

    # Estilos
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="4BB8C8", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_side = Side(style="thin", color="2A3040")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    alt_fill = PatternFill("solid", fgColor="161B22")

    headers = [
        "Folio", "Fecha Captura", "Fecha Solicitud", "Cliente", "Tipo", "Subtipo",
        "Comercial", "Ingeniero", "Prioridad", "Estatus",
        "Monto Oportunidad", "Días Captura", "Días Sin Mov.",
        "Fecha Compromiso", "Fecha Envío Cliente", "Fecha Cierre",
        "Hist. Surtido", "Inventario", "Maestro Prod.", "Hist. Recepción", "Cuestionario Log.",
        "# Rutas", "Orígenes", "Destinos", "Tipos Servicio", "Unidades",
        "Peso (kg)", "Kg/Entrega (LTL)", "M3/Entrega (LTL)", "Temperaturas", "Custodias"
    ]
    col_widths = [16,16,16,28,22,12,22,22,10,24,18,12,12,16,18,14,14,12,14,15,16,8,32,32,22,20,12,14,14,18,18]

    ws.row_dimensions[1].height = 30
    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border
        ws.column_dimensions[get_column_letter(col_num)].width = width

    for row_num, s in enumerate(solicitudes, 2):
        fill = alt_fill if row_num % 2 == 0 else PatternFill("solid", fgColor="1A1F27")
        font = Font(color="E2E8F0", size=10)
        rutas = list(s.rutas) if 'transporte' in s.tema.lower() else []
        row_data = [
            s.folio,
            s.fecha_captura.strftime('%d/%m/%Y %H:%M') if s.fecha_captura else '',
            s.fecha_solicitud.strftime('%d/%m/%Y') if s.fecha_solicitud else '',
            s.cliente,
            s.tema,
            s.subtipo or '',
            s.hunter.nombre if s.hunter else '',
            s.responsable.nombre if s.responsable else 'Sin asignar',
            s.prioridad or '',
            s.estatus,
            s.monto_oportunidad or 0,
            s.dias_desde_captura(),
            s.dias_sin_movimiento(),
            s.fecha_compromiso.strftime('%d/%m/%Y') if s.fecha_compromiso else '',
            s.fecha_envio_cliente.strftime('%d/%m/%Y') if s.fecha_envio_cliente else '',
            s.fecha_cierre.strftime('%d/%m/%Y') if s.fecha_cierre else '',
            'Sí' if s.historial_surtido else 'No',
            'Sí' if s.inventario else 'No',
            'Sí' if s.maestro_productos else 'No',
            'Sí' if s.historial_recepcion else 'No',
            'Sí' if s.cuestionario_logistico else 'No',
            len(rutas) if rutas else '',
            ' | '.join(r.origen for r in rutas),
            ' | '.join(r.destino for r in rutas),
            ' | '.join(r.tipo_servicio for r in rutas),
            ' | '.join(r.tipo_unidad or '' for r in rutas),
            ' | '.join(str(r.peso_kg or '') for r in rutas),
            ' | '.join(str(r.kg_por_entrega or '') for r in rutas),
            ' | '.join(str(r.m3_por_entrega or '') for r in rutas),
            ' | '.join(r.temperatura or '' for r in rutas),
            ' | '.join(r.custodia or '' for r in rutas),
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = font
            cell.fill = fill
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
            # Formato moneda para monto
            if col_num == 10:
                cell.number_format = '"$"#,##0'
                cell.font = Font(color="AECA00", bold=True, size=10)
            # Color estatus
            if col_num == 9:
                colores_estatus = {
                    'Capturada': 'E2E8F0', 'Asignada': '4BB8C8',
                    'En Análisis': 'F59E0B', 'Información Completa': 'AECA00',
                    'Propuesta Enviada': '7FAD00', 'Cerrada': '6B7280',
                    'Pendiente Información Cliente': 'EF4444'
                }
                color = colores_estatus.get(value, 'E2E8F0')
                cell.font = Font(color=color, bold=True, size=10)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Summary row
    total_row = len(solicitudes) + 3
    ws.cell(row=total_row, column=1, value="TOTAL SOLICITUDES").font = Font(bold=True, color="4BB8C8", size=11)
    ws.cell(row=total_row, column=2, value=len(solicitudes)).font = Font(bold=True, color="E2E8F0", size=11)
    ws.cell(row=total_row, column=9, value="MONTO TOTAL").font = Font(bold=True, color="4BB8C8", size=11)
    total_monto = sum(s.monto_oportunidad or 0 for s in solicitudes)
    monto_cell = ws.cell(row=total_row, column=10, value=total_monto)
    monto_cell.font = Font(bold=True, color="AECA00", size=11)
    monto_cell.number_format = '"$"#,##0'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fecha_str = cdmx_now().strftime('%Y%m%d_%H%M')
    filename = f"SeguimientoSD_Solicitudes_{fecha_str}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.route('/exportar/usuarios')
@login_required
@rol_requerido('administrador','lider_comercial','lider_soluciones','aux_comercial')
def exportar_usuarios():
    if es_admin():
        usuarios = User.query.order_by(User.rol, User.nombre).all()
    elif current_user.rol in ('lider_comercial','aux_comercial'):
        usuarios = User.query.filter(User.rol.in_(['comercial','aux_comercial','lider_comercial']))                             .order_by(User.nombre).all()
    else:
        usuarios = User.query.filter(User.rol.in_(['ingeniero','aux_ingenieria','lider_soluciones']))                             .order_by(User.nombre).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Nombre', 'Usuario', 'Rol', 'Estatus', 'Fecha Creación',
                     'Solicitudes Creadas', 'Solicitudes Asignadas'])
    for u in usuarios:
        writer.writerow([
            u.nombre,
            u.username,
            ROLES_LABEL.get(u.rol, u.rol),
            'Activo' if u.activo else 'Inactivo',
            u.created_at.strftime('%d/%m/%Y') if u.created_at else '',
            u.solicitudes_creadas.count(),
            u.solicitudes_asignadas.count(),
        ])

    fecha_str = cdmx_now().strftime('%Y%m%d_%H%M')
    filename = f"SeguimientoSD_Usuarios_{fecha_str}.csv"
    return Response(
        '﻿' + output.getvalue(),   # BOM para que Excel lo abra bien
        mimetype='text/csv; charset=utf-8',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ══════════════════════════════════════════════════════════════════════════════
# ── RESPALDO DE ARCHIVOS ─────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/admin/backup')
@login_required
@rol_requerido('administrador')
def backup_uploads():
    """Página de respaldo — muestra estadísticas antes de descargar."""
    stats = {'total_archivos': 0, 'total_size': 0, 'solicitudes': []}

    if os.path.exists(UPLOAD_BASE):
        for folio in sorted(os.listdir(UPLOAD_BASE)):
            folio_path = os.path.join(UPLOAD_BASE, folio)
            if not os.path.isdir(folio_path):
                continue
            archivos = []
            for tipo in ['comercial', 'soluciones']:
                tipo_path = os.path.join(folio_path, tipo)
                if os.path.isdir(tipo_path):
                    for fname in os.listdir(tipo_path):
                        fpath = os.path.join(tipo_path, fname)
                        if os.path.isfile(fpath):
                            size = os.path.getsize(fpath)
                            # Buscar nombre original en BD
                            doc = Documento.query.filter_by(
                                nombre_guardado=fname, activo=True).first()
                            archivos.append({
                                'tipo': tipo,
                                'nombre': doc.nombre_original if doc else fname,
                                'size': size,
                            })
                            stats['total_archivos'] += 1
                            stats['total_size'] += size
            if archivos:
                stats['solicitudes'].append({
                    'folio': folio,
                    'archivos': archivos,
                    'count': len(archivos),
                })

    stats['total_size_mb'] = round(stats['total_size'] / 1024 / 1024, 2)
    return render_template('admin_backup.html', stats=stats)


@app.route('/admin/backup/descargar')
@login_required
@rol_requerido('administrador')
def descargar_backup():
    """Genera y descarga un ZIP con todos los archivos de uploads/."""
    buf = io.BytesIO()

    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        if os.path.exists(UPLOAD_BASE):
            for folio in os.listdir(UPLOAD_BASE):
                folio_path = os.path.join(UPLOAD_BASE, folio)
                if not os.path.isdir(folio_path):
                    continue
                for tipo in ['comercial', 'soluciones']:
                    tipo_path = os.path.join(folio_path, tipo)
                    if not os.path.isdir(tipo_path):
                        continue
                    for fname in os.listdir(tipo_path):
                        fpath = os.path.join(tipo_path, fname)
                        if not os.path.isfile(fpath):
                            continue
                        # Buscar nombre original para el ZIP
                        doc = Documento.query.filter_by(
                            nombre_guardado=fname, activo=True).first()
                        nombre_zip = doc.nombre_original if doc else fname
                        # Ruta dentro del ZIP: folio/tipo/nombre_original
                        arcname = os.path.join(folio, tipo, nombre_zip)
                        zf.write(fpath, arcname)

    buf.seek(0)
    fecha_str = cdmx_now().strftime('%Y%m%d_%H%M')
    filename = f'SeguimientoSD_Archivos_{fecha_str}.zip'
    return Response(
        buf.getvalue(),
        mimetype='application/zip',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


# ── Migración temporal — ELIMINAR después de correr una vez ──────────────────
@app.route('/admin/migrar-columnas')
@login_required
@rol_requerido('administrador')
def migrar_columnas():
    try:
        with db.engine.connect() as conn:
            from sqlalchemy import text
            conn.execute(text("ALTER TABLE solicitudes ADD COLUMN IF NOT EXISTS subtipo VARCHAR(10)"))
            conn.execute(text("ALTER TABLE solicitudes ADD COLUMN IF NOT EXISTS prioridad_sugerida INTEGER"))
            conn.execute(text("ALTER TABLE solicitudes ADD COLUMN IF NOT EXISTS prioridad_comercial INTEGER"))
            conn.execute(text("ALTER TABLE solicitudes ADD COLUMN IF NOT EXISTS prioridad_estado VARCHAR(20) DEFAULT 'pendiente'"))
            conn.commit()
        flash("✅ Migración completada. Ya puedes eliminar esta ruta del código.", "success")
    except Exception as e:
        flash(f"Error en migración: {str(e)}", "danger")
    return redirect(url_for("dashboard"))


# ── Init DB ───────────────────────────────────────────────────────────────────
def init_db():
    db.create_all()
    if TemasSolicitud.query.count() == 0:
        for i, nombre in enumerate(TEMAS_DEFAULT):
            db.session.add(TemasSolicitud(nombre=nombre, orden=i))
        db.session.commit()
        print('Temas inicializados.')
    if User.query.count() == 0:
        db.session.add(User(
            username='admin',
            password_hash=generate_password_hash('Admin2026!1'),
            nombre='Administrador', rol='administrador', activo=True
        ))
        db.session.commit()
        print('Admin creado. Ejecuta /admin/reset-passwords para cargar usuarios.')




@app.route('/admin/reset-passwords')
@login_required
@rol_requerido('administrador')
def reset_passwords():
    from werkzeug.security import generate_password_hash as gph
    USUARIOS = [
        ('Administrador',       'admin',                'administrador',    True,  'Admin2026!1'),
        ('Rubi Arizmendi',      'Rubi_Arizmendi',       'aux_comercial',    True,  'AuxCom2026!1'),
        ('Aline Esquivel',      'Aline_Esquivel',       'aux_ingenieria',   True,  'AuxCom2026!2'),
        ('Teresa Ruiz',         'Teresa_Ruiz',          'comercial',        True,  'Hunter2026!1'),
        ('Alejandra Sanchez',   'Alejandra_Sanchez',    'comercial',        True,  'Hunter2026!2'),
        ('Diana Pelcastre',     'Diana_Pelcastre',      'comercial',        True,  'Hunter2026!3'),
        ('Ida Acosta',          'Ida_Acosta',           'comercial',        True,  'Hunter2026!4'),
        ('Malena Baltazar',     'Malena_Baltazar',      'comercial',        True,  'Hunter2026!5'),
        ('Jose Ortega',         'Jose_Ortega',          'comercial',        True,  'Hunter2026!6'),
        ('Karla Herrera',       'Karla_Herrera',        'comercial',        True,  'Hunter2026!7'),
        ('David Fortoul',       'David_Fortoul',        'comercial',        True,  'Hunter2026!8'),
        ('Maria Elena Baltazar','Maria_Elena_Baltazar', 'comercial',        True,  'Hunter2026!9'),
        ('Genoveva Roa',        'Genoveva_Roa',         'comercial',        True,  'Hunter2026!10'),
        ('Elizabeth Bastida',   'Elizabeth_Bastida',    'ingeniero',        True,  'IngeSD2026!1'),
        ('Diego Arzate',        'Diego_Arzate',         'ingeniero',        True,  'IngeSD2026!2'),
        ('Jorge Camarena',      'Jorge_Camarena',       'ingeniero',        True,  'IngeSD2026!3'),
        ('Jose Luis Montes',    'Jose_Luis_Montes',     'ingeniero',        False, 'IngeSD2026!4'),
        ('Gerardo Velazquez',   'Gerardo_Velazquez',    'ingeniero',        True,  'IngeSD2026!5'),
        ('Francisco Cueva',     'Francisco_Cueva',      'lider_comercial',  True,  'Lcomercial1231'),
        ('Andres Toledo',       'Andres_Toledo',        'lider_soluciones', True,  'Lsoluciones1231'),
    ]
    log = []
    for nombre, username, rol, activo, pwd in USUARIOS:
        u = User.query.filter(db.or_(User.nombre==nombre, User.username==username)).first()
        if u:
            u.nombre=nombre; u.username=username; u.rol=rol
            u.activo=activo; u.password_hash=gph(pwd)
            log.append('UPD|'+nombre+'|'+username+'|'+rol+'|'+pwd)
        else:
            db.session.add(User(nombre=nombre,username=username,rol=rol,
                activo=activo,password_hash=gph(pwd)))
            log.append('NEW|'+nombre+'|'+username+'|'+rol+'|'+pwd)
    db.session.commit()
    rows=''.join('<tr>'+''.join('<td style="padding:6px 10px;border-bottom:1px solid #2a3040">'+c+'</td>' for c in l.split('|'))+'</tr>' for l in log)
    return ('<html><head><meta charset="UTF-8"><style>body{font-family:Arial;background:#0f172a;color:#e2e8f0;padding:30px}'
            'table{border-collapse:collapse;width:100%;font-size:12px}th{color:#4BB8C8;padding:8px 10px;border-bottom:2px solid #4BB8C8;text-align:left}'
            'h2{color:#4BB8C8}a{color:#4BB8C8;margin-right:12px}</style></head><body>'
            '<h2>Usuarios sincronizados</h2><table><tr><th>Accion</th><th>Nombre</th><th>Username</th><th>Rol</th><th>Password</th></tr>'
            +rows+'</table><br><a href="/admin/carga-v2">Cargar solicitudes</a><a href="/dashboard">Dashboard</a></body></html>')


@app.route('/admin/carga-v2')
@login_required
@rol_requerido('administrador')
def carga_v2():
    from openpyxl import load_workbook as lw
    from datetime import datetime as dt2, date as date2
    paso = request.args.get('paso', 'menu')
    lote = int(request.args.get('lote', 0))
    TAM  = 20
    CSS  = ('<style>body{font-family:Arial;background:#0f172a;color:#e2e8f0;padding:30px;max-width:900px}'
            '.btn{background:#4BB8C8;color:#000;border:none;padding:10px 22px;font-size:14px;font-weight:700;'
            'border-radius:6px;cursor:pointer;text-decoration:none;display:inline-block;margin:6px}'
            '.g{background:#7fad00;color:#000}'
            'pre{background:#161b22;border:1px solid #2a3040;border-radius:8px;padding:16px;font-size:12px;'
            'max-height:420px;overflow:auto;white-space:pre-wrap}'
            '.card{background:#161b22;border:1px solid #2a3040;border-radius:8px;padding:20px;margin:12px 0}'
            'h2{color:#4BB8C8}h3{color:#aeca00}</style>')

    if paso == 'menu':
        tu = User.query.count()
        ts = Solicitud.query.count()
        return (CSS+'<h2>Carga Masiva</h2>'
                '<div class="card">Estado: <b>'+str(tu)+' usuarios</b> · <b>'+str(ts)+' solicitudes</b></div>'
                '<div class="card"><h3>Paso 1 - Usuarios</h3>'
                '<a class="btn" href="/admin/reset-passwords">Sincronizar usuarios</a></div>'
                '<div class="card"><h3>Paso 2 - Solicitudes (lotes de 20)</h3>'
                '<a class="btn" href="/admin/carga-v2?paso=solicitudes&lote=0">Iniciar carga</a></div>')

    if paso == 'solicitudes':
        EMAP = {'Entregado':'Cerrada','Pendiente':'Pendiente de Informacion',
                'On Hold':'Pendiente de Informacion','En proceso':'En Proceso'}

        def pf(v):
            if not v: return None
            if isinstance(v, dt2): return v.date()
            if isinstance(v, date2): return v
            for fmt in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y'):
                try: return dt2.strptime(str(v).strip(), fmt).date()
                except: pass
            return None

        def ck(v):
            return bool(v) and str(v).strip() not in ('','- ','None','nan')

        wb = lw('/opt/render/project/src/carga_inicial_data.xlsx', data_only=True)
        ws = wb['Soluciones']
        all_rows = []
        for row in range(2, ws.max_row+1):
            vals = [ws.cell(row,c).value for c in range(1,26)]
            if any(vals[:5]):
                all_rows.append(vals)

        total_lotes = (len(all_rows)+TAM-1)//TAM
        inicio = lote*TAM; fin = min(inicio+TAM, len(all_rows))
        bloque = all_rows[inicio:fin]

        if not bloque:
            ts = Solicitud.query.count()
            return CSS+'<h2>Carga completada!</h2><div class="card">Total: <b>'+str(ts)+'</b></div><a class="btn g" href="/dashboard">Dashboard</a>'

        ubn     = {u.nombre: u for u in User.query.all()}
        admin_u = User.query.filter_by(username='admin').first()

        def nuevo_folio():
            ultima = Solicitud.query.filter(Solicitud.folio.like('SOL-2026-%')).order_by(Solicitud.id.desc()).first()
            num = int(ultima.folio.split('-')[-1])+1 if ultima else 1
            return 'SOL-2026-'+str(num).zfill(4)

        log = ['Lote '+str(lote+1)+'/'+str(total_lotes)+' (filas '+str(inicio+2)+'-'+str(fin+1)+')']
        ok  = 0

        for r in bloque:
            cliente = str(r[2]).strip() if r[2] else ''
            if not cliente: continue
            hunter  = ubn.get(str(r[0]).strip() if r[0] else '', admin_u) or admin_u
            resp_n  = str(r[3]).strip() if r[3] else ''
            resp    = ubn.get(resp_n) if resp_n and resp_n.lower() != 'sin asignar' else None
            status_r= str(r[9]).strip() if r[9] else ''
            estatus = EMAP.get(status_r, 'Asignada')
            f_sol   = pf(r[1]); f_comp = pf(r[5]); f_ent = pf(r[8])
            ahora   = dt2.now()
            f_cap   = dt2.combine(f_sol, dt2.min.time()) if f_sol else ahora

            sol = Solicitud(
                folio=nuevo_folio(), hunter_id=hunter.id,
                responsable_id=resp.id if resp else None,
                fecha_solicitud=f_sol or date2(2026,1,1),
                fecha_captura=f_cap, ultima_actualizacion=f_cap,
                cliente=cliente, tema=str(r[16]).strip() if r[16] else 'Sin definir',
                comentarios_comerciales=str(r[24]).strip() if r[24] else '',
                estatus=estatus, fecha_compromiso=f_comp,
                cuestionario_logistico=ck(r[18]), inventario=ck(r[22]),
                maestro_productos=ck(r[23]), prioridad=None, prioridad_estado='pendiente',
            )
            if estatus == 'Cerrada':
                c = dt2.combine(f_ent, dt2.min.time()) if f_ent else ahora
                sol.fecha_cierre=c; sol.usuario_cierre_id=admin_u.id
                sol.fecha_envio_cliente=c; sol.usuario_envio_id=admin_u.id
                sol.ultima_actualizacion=c

            db.session.add(sol); db.session.flush()
            db.session.add(Bitacora(solicitud_id=sol.id, usuario_id=admin_u.id,
                accion='Carga masiva. Status: '+status_r))
            if resp:
                db.session.add(SolicitudIngeniero(
                    solicitud_id=sol.id, ingeniero_id=resp.id, es_principal=True))
            log.append('OK '+sol.folio+' | '+cliente[:35]+' | '+estatus)
            ok += 1

        db.session.commit()
        ts  = Solicitud.query.count()
        sig = lote+1
        log.append(str(ok)+' insertadas. Total BD: '+str(ts)+'/'+str(len(all_rows)))
        btn = ('<a class="btn g" href="/admin/carga-v2?paso=solicitudes&lote='+str(sig)+'">Siguiente lote ('+str(sig+1)+'/'+str(total_lotes)+') -></a>'
               if fin < len(all_rows) else
               '<a class="btn g" href="/admin/carga-v2?paso=solicitudes&lote='+str(sig)+'">Ver resumen</a>')
        salida  = CSS+'<h2>Lote '+str(lote+1)+'/'+str(total_lotes)+'</h2>'
        salida += '<pre>'+chr(10).join(log)+'</pre>'+btn
        salida += ' <a class="btn" href="/admin/carga-v2">Menu</a>'
        return salida

    return 'Paso no valido', 400



@app.route('/admin/fix-duplicados')
@login_required
@rol_requerido('administrador')
def fix_duplicados():
    """Elimina solicitudes duplicadas manteniendo la primera ocurrencia por cliente+tema+fecha."""
    from sqlalchemy import func
    
    # Find duplicates - keep lowest id per cliente+tema+fecha_solicitud
    todas = Solicitud.query.order_by(Solicitud.id.asc()).all()
    
    vistos = {}
    a_eliminar = []
    for s in todas:
        key = (s.cliente.strip().lower(), s.tema.strip().lower(), str(s.fecha_solicitud))
        if key in vistos:
            a_eliminar.append(s)
        else:
            vistos[key] = s.id

    eliminadas = 0
    for s in a_eliminar:
        Bitacora.query.filter_by(solicitud_id=s.id).delete()
        SolicitudIngeniero.query.filter_by(solicitud_id=s.id).delete()
        db.session.delete(s)
        eliminadas += 1

    db.session.commit()
    total = Solicitud.query.count()
    return ('<html><head><meta charset="UTF-8"><style>'
            'body{font-family:Arial;background:#0f172a;color:#e2e8f0;padding:30px}'
            'h2{color:#4BB8C8}a{color:#4BB8C8}</style></head><body>'
            '<h2>Duplicados eliminados</h2>'
            '<p>Eliminadas: <b>' + str(eliminadas) + '</b> solicitudes duplicadas.</p>'
            '<p>Total restante: <b>' + str(total) + '</b> solicitudes.</p>'
            '<a href="/dashboard">Ir al Dashboard</a>'
            '</body></html>')


@app.route('/admin/fix-usuarios-duplicados')
@login_required
@rol_requerido('administrador')
def fix_usuarios_duplicados():
    """Elimina usuarios duplicados manteniendo el de username correcto (con acentos/mayúsculas)."""
    from werkzeug.security import generate_password_hash as gph

    # Usuarios definitivos — el username correcto gana
    DEFINITIVOS = {
        'admin':                ('Administrador',       'administrador',    True,  'Admin2026!1'),
        'Rubi_Arizmendi':       ('Rubi Arizmendi',      'aux_comercial',    True,  'AuxCom2026!1'),
        'Aline_Esquivel':       ('Aline Esquivel',      'aux_ingenieria',   True,  'AuxCom2026!2'),
        'Teresa_Ruiz':          ('Teresa Ruiz',         'comercial',        True,  'Hunter2026!1'),
        'Alejandra_Sanchez':    ('Alejandra Sanchez',   'comercial',        True,  'Hunter2026!2'),
        'Diana_Pelcastre':      ('Diana Pelcastre',     'comercial',        True,  'Hunter2026!3'),
        'Ida_Acosta':           ('Ida Acosta',          'comercial',        True,  'Hunter2026!4'),
        'Malena_Baltazar':      ('Malena Baltazar',     'comercial',        True,  'Hunter2026!5'),
        'Jose_Ortega':          ('Jose Ortega',         'comercial',        True,  'Hunter2026!6'),
        'Karla_Herrera':        ('Karla Herrera',       'comercial',        True,  'Hunter2026!7'),
        'David_Fortoul':        ('David Fortoul',       'comercial',        True,  'Hunter2026!8'),
        'Maria_Elena_Baltazar': ('Maria Elena Baltazar','comercial',        True,  'Hunter2026!9'),
        'Genoveva_Roa':         ('Genoveva Roa',        'comercial',        True,  'Hunter2026!10'),
        'Elizabeth_Bastida':    ('Elizabeth Bastida',   'ingeniero',        True,  'IngeSD2026!1'),
        'Diego_Arzate':         ('Diego Arzate',        'ingeniero',        True,  'IngeSD2026!2'),
        'Jorge_Camarena':       ('Jorge Camarena',      'ingeniero',        True,  'IngeSD2026!3'),
        'Jose_Luis_Montes':     ('Jose Luis Montes',    'ingeniero',        False, 'IngeSD2026!4'),
        'Gerardo_Velazquez':    ('Gerardo Velazquez',   'ingeniero',        True,  'IngeSD2026!5'),
        'Francisco_Cueva':      ('Francisco Cueva',     'lider_comercial',  True,  'Lcomercial1231'),
        'Andres_Toledo':        ('Andres Toledo',       'lider_soluciones', True,  'Lsoluciones1231'),
    }

    todos = User.query.all()
    log = []
    eliminados = 0
    actualizados = 0

    # Find users NOT in definitivos list (old/duplicate usernames)
    usernames_validos = set(DEFINITIVOS.keys())
    for u in todos:
        if u.username not in usernames_validos:
            # Check if there is a solicitud with this user as hunter or responsable
            tiene_sols = (Solicitud.query.filter(
                db.or_(Solicitud.hunter_id==u.id, Solicitud.responsable_id==u.id)
            ).count() > 0)

            if tiene_sols:
                # Reassign solicitudes to correct user before deleting
                # Find correct user by similar nombre
                nombre_lower = u.nombre.lower().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('ñ','n')
                correcto = None
                for uname, (nombre, rol, activo, pwd) in DEFINITIVOS.items():
                    nombre_def = nombre.lower()
                    if nombre_lower == nombre_def:
                        correcto = User.query.filter_by(username=uname).first()
                        break

                if correcto:
                    Solicitud.query.filter_by(hunter_id=u.id).update({'hunter_id': correcto.id})
                    Solicitud.query.filter_by(responsable_id=u.id).update({'responsable_id': correcto.id})
                    SolicitudIngeniero.query.filter_by(ingeniero_id=u.id).update({'ingeniero_id': correcto.id})
                    Bitacora.query.filter_by(usuario_id=u.id).update({'usuario_id': correcto.id})
                    log.append('REASIGNADO ' + u.username + ' -> ' + correcto.username)
                else:
                    log.append('SIN MATCH para ' + u.username + ' (' + u.nombre + ')')
                    continue

            db.session.delete(u)
            eliminados += 1
            log.append('ELIMINADO ' + u.username)

    # Update all valid users
    for username, (nombre, rol, activo, pwd) in DEFINITIVOS.items():
        u = User.query.filter_by(username=username).first()
        if u:
            u.nombre = nombre; u.rol = rol; u.activo = activo
            u.password_hash = gph(pwd)
            actualizados += 1

    db.session.commit()

    total = User.query.count()
    rows = ''.join('<tr><td style="padding:5px 10px;border-bottom:1px solid #2a3040">' + l + '</td></tr>' for l in log)
    return ('<html><head><meta charset="UTF-8"><style>'
            'body{font-family:Arial;background:#0f172a;color:#e2e8f0;padding:30px}'
            'table{border-collapse:collapse;width:100%;font-size:12px}'
            'h2{color:#4BB8C8}a{color:#4BB8C8}</style></head><body>'
            '<h2>Usuarios limpiados</h2>'
            '<p>Eliminados: <b>' + str(eliminados) + '</b> | '
            'Actualizados: <b>' + str(actualizados) + '</b> | '
            'Total actual: <b>' + str(total) + '</b></p>'
            '<table>' + rows + '</table>'
            '<br><a href="/admin/usuarios">Ver usuarios</a> '
            '<a href="/dashboard">Dashboard</a>'
            '</body></html>')

if __name__ == '__main__':
    port  = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(host='0.0.0.0', port=port, debug=debug)
